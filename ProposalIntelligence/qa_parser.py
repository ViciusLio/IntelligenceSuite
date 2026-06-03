"""Parser Q&A per formati strutturati.

Due responsabilità:
  * ``parse_qa_pairs(path)``  → coppie (domanda, risposta) dal **corpus di stile**
  * ``parse_questions(path)`` → sole domande da un **questionario in arrivo**

Formati supportati (v1, strutturati):
  * Markdown / CSV: tabella a 2 colonne (Domanda | Risposta)
  * Markdown / TXT: marcatori espliciti  D:/R:  ·  Q:/A:  ·  Domanda:/Risposta:
  * Excel (.xlsx): prime due colonne (richiede l'extra ``[xlsx]``)
  * Solo-domande: liste numerate/puntate, colonna singola, righe con "?"

PDF/DOCX free-form sono volutamente fuori scope in v1: il pairing Q↔A su testo
non strutturato è euristico e poco affidabile.
"""

from __future__ import annotations

import csv
import re
from pathlib import Path

# Marcatori di domanda/risposta accettati (case-insensitive), IT + EN.
_Q_MARKERS = ("domanda", "question", "d", "q")
_A_MARKERS = ("risposta", "answer", "r", "a")

_QA_MARKER_RE = re.compile(
    r"^\s*(?P<tag>[A-Za-zÀ-ÿ]+)\s*[:\.\)]\s*(?P<body>.*)$"
)
_NUMBERED_RE = re.compile(r"^\s*\d+\s*[\.\)]\s+(?P<body>.+?)\s*$")
_BULLET_RE = re.compile(r"^\s*[-*•]\s+(?P<body>.+?)\s*$")

_HTML_COMMENT_RE = re.compile(r"<!--.*?-->", re.DOTALL)


# ── helpers ────────────────────────────────────────────────────────────────
def _norm(s: str) -> str:
    return re.sub(r"\s+", " ", (s or "").strip())


def _strip_html_comments(text: str) -> str:
    """Rimuove i blocchi ``<!-- ... -->`` (anche multi-riga) prima del parsing."""
    return _HTML_COMMENT_RE.sub("", text)


def _is_table_separator(line: str) -> bool:
    """Riga separatrice di una tabella Markdown:  |---|:--:|"""
    stripped = line.strip().strip("|")
    return bool(stripped) and set(stripped) <= set("-:| \t")


def _split_table_row(line: str) -> list[str]:
    cells = line.strip().strip("|").split("|")
    return [_norm(c) for c in cells]


def _looks_like_header(q_cell: str) -> bool:
    head = q_cell.lower().strip()
    return head in {"domanda", "domande", "question", "questions", "d", "q", "quesito"}


# ── corpus: coppie Domanda/Risposta ─────────────────────────────────────────
def parse_qa_pairs(path: Path) -> list[tuple[str, str]]:
    """Estrae coppie (domanda, risposta) dal file corpus."""
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return _pairs_from_csv(path)
    if suffix == ".xlsx":
        return _pairs_from_xlsx(path)
    # .md / .markdown / .txt
    text = _strip_html_comments(path.read_text(encoding="utf-8"))
    pairs = _pairs_from_markdown_table(text)
    if pairs:
        return pairs
    return _pairs_from_markers(text)


def _pairs_from_markdown_table(text: str) -> list[tuple[str, str]]:
    pairs: list[tuple[str, str]] = []
    rows = [
        ln for ln in text.splitlines()
        if "|" in ln and not _is_table_separator(ln)
    ]
    for ln in rows:
        cells = _split_table_row(ln)
        if len(cells) < 2:
            continue
        q, a = cells[0], cells[1]
        if _looks_like_header(q):
            continue   # salta intestazione (in qualunque posizione)
        if q and a:
            pairs.append((q, a))
    return pairs


def _pairs_from_markers(text: str) -> list[tuple[str, str]]:
    """Coppie da marcatori D:/R: (o Q:/A:, Domanda:/Risposta:).

    Gestisce risposte multi-riga: accumula finché non incontra il marcatore
    successivo o una riga vuota seguita da una nuova domanda.
    """
    pairs: list[tuple[str, str]] = []
    cur_q: str | None = None
    cur_a: list[str] = []
    mode: str | None = None  # "q" | "a"

    def _flush() -> None:
        nonlocal cur_q, cur_a, mode
        if cur_q and cur_a:
            ans = _norm(" ".join(cur_a))
            if ans:
                pairs.append((_norm(cur_q), ans))
        cur_q, cur_a, mode = None, [], None

    for raw in text.splitlines():
        m = _QA_MARKER_RE.match(raw)
        tag = m.group("tag").lower() if m else None
        if tag in _Q_MARKERS:
            _flush()
            cur_q, mode = m.group("body"), "q"
        elif tag in _A_MARKERS and cur_q is not None:
            mode = "a"
            if m.group("body"):
                cur_a.append(m.group("body"))
        else:
            # continuazione della riga corrente (domanda o risposta multi-riga)
            if mode == "a" and raw.strip():
                cur_a.append(raw.strip())
            elif mode == "q" and raw.strip():
                cur_q = f"{cur_q} {raw.strip()}"
    _flush()
    return pairs


def _pairs_from_csv(path: Path) -> list[tuple[str, str]]:
    pairs: list[tuple[str, str]] = []
    with path.open(encoding="utf-8", newline="") as f:
        reader = csv.reader(f)
        for idx, row in enumerate(reader):
            if len(row) < 2:
                continue
            q, a = _norm(row[0]), _norm(row[1])
            if idx == 0 and _looks_like_header(q):
                continue
            if q and a:
                pairs.append((q, a))
    return pairs


def _pairs_from_xlsx(path: Path) -> list[tuple[str, str]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - dipende dall'extra
        raise ImportError(
            "Lettura .xlsx richiede openpyxl: pip install 'intelligence-suite[xlsx]'"
        ) from exc
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    pairs: list[tuple[str, str]] = []
    for idx, row in enumerate(ws.iter_rows(values_only=True)):
        if not row or len(row) < 2:
            continue
        q, a = _norm(str(row[0] or "")), _norm(str(row[1] or ""))
        if idx == 0 and _looks_like_header(q):
            continue
        if q and a:
            pairs.append((q, a))
    wb.close()
    return pairs


# ── questionario in arrivo: sole domande ─────────────────────────────────────
def parse_questions(path: Path) -> list[str]:
    """Estrae l'elenco di domande da un questionario (senza risposte)."""
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return _questions_from_rows(_csv_first_column(path))
    if suffix == ".xlsx":
        return _questions_from_rows(_xlsx_first_column(path))
    return _questions_from_text(_strip_html_comments(path.read_text(encoding="utf-8")))


def _questions_from_text(text: str) -> list[str]:
    # 1) tabella Markdown → prima colonna
    table_rows = [
        ln for ln in text.splitlines()
        if "|" in ln and not _is_table_separator(ln)
    ]
    if table_rows:
        out: list[str] = []
        for ln in table_rows:
            cells = _split_table_row(ln)
            if not cells or not cells[0]:
                continue
            if _looks_like_header(cells[0]):
                continue
            out.append(cells[0])
        if out:
            return out

    # 2) liste numerate / puntate / marcatori D: / righe con "?"
    questions: list[str] = []
    for raw in text.splitlines():
        line = raw.strip()
        if not line:
            continue
        m = _NUMBERED_RE.match(line) or _BULLET_RE.match(line)
        if m:
            questions.append(_norm(m.group("body")))
            continue
        mk = _QA_MARKER_RE.match(line)
        if mk and mk.group("tag").lower() in _Q_MARKERS and mk.group("body"):
            questions.append(_norm(mk.group("body")))
            continue
        if line.endswith("?"):
            questions.append(_norm(line))
    return questions


def _questions_from_rows(cells: list[str]) -> list[str]:
    out = []
    for idx, c in enumerate(cells):
        c = _norm(c)
        if not c:
            continue
        if idx == 0 and _looks_like_header(c):
            continue
        out.append(c)
    return out


def _csv_first_column(path: Path) -> list[str]:
    with path.open(encoding="utf-8", newline="") as f:
        return [row[0] for row in csv.reader(f) if row]


def _xlsx_first_column(path: Path) -> list[str]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover
        raise ImportError(
            "Lettura .xlsx richiede openpyxl: pip install 'intelligence-suite[xlsx]'"
        ) from exc
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    cells = [str(row[0]) for row in ws.iter_rows(values_only=True) if row and row[0]]
    wb.close()
    return cells
