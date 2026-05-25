"""Parser XLSX — chunk per foglio e per regione di dati (tabelle)."""

from __future__ import annotations
import re
from pathlib import Path

from intelligence_core.chunk import make_chunk

try:
    import openpyxl
    HAS_XLSX = True
except ImportError:
    HAS_XLSX = False


def can_parse(path: Path) -> bool:
    return path.suffix.lower() in {".xlsx", ".xlsm", ".xltx"}


def parse_file(path: Path, root: Path) -> list[dict]:
    """Ritorna chunk da XLSX. Fallback a chunk raw."""
    rel = str(path.relative_to(root)).replace("\\", "/")
    stem = path.stem

    if not HAS_XLSX:
        return [make_chunk(
            domain="doc", type_="section", locator=stem,
            text=f"File: {path.name} (in {rel})\n---\n[openpyxl non installato]",
            source=rel, language="xlsx",
            metadata={"page_start": 1, "page_end": 1, "heading_path": [stem]},
        )]

    try:
        return _parse_xlsx(path, rel, stem)
    except Exception as e:
        return [make_chunk(
            domain="doc", type_="section", locator=stem,
            text=f"File: {path.name} (in {rel})\n---\n[Errore parsing: {e}]",
            source=rel, language="xlsx",
            metadata={"page_start": 1, "page_end": 1, "heading_path": [stem]},
        )]


def _parse_xlsx(path: Path, rel: str, stem: str) -> list[dict]:
    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    chunks = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))

        # Filtra righe completamente vuote
        rows = [r for r in rows if any(v is not None and str(v).strip() for v in r)]
        if not rows:
            continue

        # Chunk sezione per il foglio
        sheet_text = (
            f"Section: {sheet_name} (in {path.name})\n"
            f"Path: {stem} > {sheet_name}\n"
            f"---\n"
            f"Foglio con {len(rows)} righe."
        )
        chunks.append(make_chunk(
            domain="doc", type_="section",
            locator=f"{stem}.{_sanitize(sheet_name)}",
            text=sheet_text, source=rel, language="xlsx",
            metadata={
                "page_start": 1, "page_end": 1,
                "heading_path": [stem, sheet_name],
                "sheet": sheet_name,
                "row_count": len(rows),
            },
        ))

        # Chunk tabella per la regione di dati
        if len(rows) >= 2:
            header = [str(v) if v is not None else "" for v in rows[0]]
            data_rows = rows[1:]
            total_rows = len(data_rows)

            preview_rows = data_rows[:20]
            lines = [
                f"Table: {sheet_name} (in {path.name})",
                f"Columns: {', '.join(header)}",
                "---",
                " | ".join(header),
            ]
            for row in preview_rows:
                lines.append(" | ".join(str(v) if v is not None else "" for v in row))
            if total_rows > 20:
                lines.append(f"... [{total_rows} righe totali]")

            table_text = "\n".join(lines)
            chunks.append(make_chunk(
                domain="doc", type_="table",
                locator=f"{stem}.{_sanitize(sheet_name)}.table",
                text=table_text[:7500], source=rel, language="xlsx",
                metadata={
                    "page_start": 1, "page_end": 1,
                    "heading_path": [stem, sheet_name, "Table"],
                    "sheet": sheet_name,
                    "columns": header,
                    "row_count": total_rows,
                },
            ))

    wb.close()
    return chunks if chunks else [make_chunk(
        domain="doc", type_="section", locator=stem,
        text=f"File: {path.name} (in {rel})\n---\n[Nessun dato trovato]",
        source=rel, language="xlsx",
        metadata={"page_start": 1, "page_end": 1, "heading_path": [stem]},
    )]


def _sanitize(s: str) -> str:
    return re.sub(r"[^\w]", "_", s.lower())[:40]
